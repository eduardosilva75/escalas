#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
import sqlite3
import pandas as pd
import threading
from PyQt5.QtCore import QTimer
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QTableWidget, QTableWidgetItem, QPushButton, QLabel,
                             QDateEdit, QSpinBox, QTabWidget, QTextEdit,
                             QMessageBox, QHeaderView, QProgressBar,
                             QFileDialog)
from PyQt5.QtCore import Qt, QDate, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QColor

class DatabaseManager:
    def __init__(self, db_path='escala_trabalho.db'):
        self.db_path = db_path
        self.init_database()

    def init_database(self):
        """Inicializa a base de dados com as tabelas necessárias e garante colunas opcionais"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Tabela de pessoas
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS pessoas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT UNIQUE NOT NULL,
                horas_diarias INTEGER DEFAULT 8,
                ativo INTEGER DEFAULT 1
            )
        ''')

        # Adicionar coluna 'cor' se não existir
        cursor.execute("PRAGMA table_info(pessoas)")
        columns = [col[1] for col in cursor.fetchall()]
        if 'cor' not in columns:
            cursor.execute('ALTER TABLE pessoas ADD COLUMN cor TEXT DEFAULT "FFFFFF"')

        # Tabela de ferias
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ferias (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pessoa_id INTEGER,
                data_inicio DATE NOT NULL,
                data_fim DATE NOT NULL,
                FOREIGN KEY (pessoa_id) REFERENCES pessoas (id)
            )
        ''')

        # Tabela de ciclos de folgas
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS folgas_ciclo (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pessoa_id INTEGER,
                semana_id INTEGER,
                dia_semana INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (pessoa_id) REFERENCES pessoas (id)
            )
        ''')

        # Tabela de horários fixos
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS horarios_fixos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pessoa_id INTEGER,
                data DATE NOT NULL,
                horario TEXT NOT NULL,
                FOREIGN KEY (pessoa_id) REFERENCES pessoas (id)
            )
        ''')

        # === TABELA UNIFICADA: dias_loja_fechada ===
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS dias_loja_fechada (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                data DATE UNIQUE NOT NULL,
                descricao TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # === MIGRAÇÃO: de loja_fechada → dias_loja_fechada ===
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='loja_fechada'")
        if cursor.fetchone():
            # Copia dados
            cursor.execute('''
                INSERT OR IGNORE INTO dias_loja_fechada (data)
                SELECT data FROM loja_fechada
            ''')
            # Apaga a tabela antiga
            cursor.execute('DROP TABLE loja_fechada')

        conn.commit()
        conn.close()

    def get_pessoas(self):
        """Obtém a lista de pessoas da base de dados com cor_hex"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        # Busca nome, horas e cor_hex
        cursor.execute('''
            SELECT p.nome, p.horas_diarias, p.cor
            FROM pessoas p
            WHERE p.ativo = 1
        ''')
        pessoas_data = cursor.fetchall()
        conn.close()

        pessoas = {}
        for nome, horas, cor_hex in pessoas_data:
            # Se cor_hex for None, usa branco
            cor = cor_hex.replace('#', '') if cor_hex and cor_hex.startswith('#') else (cor_hex if cor_hex else 'FFFFFF')
            pessoas[nome] = {'horas': horas, 'cor': cor, 'id': None}  # id não usado aqui
        return pessoas  
       

    def get_ferias(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT p.nome, f.data_inicio, f.data_fim
            FROM ferias f
            JOIN pessoas p ON f.pessoa_id = p.id
            WHERE p.ativo = 1
        ''')
        ferias_data = cursor.fetchall()
        conn.close()

        ferias = {}
        for nome, inicio, fim in ferias_data:
            if nome not in ferias:
                ferias[nome] = []
            ferias[nome].append((datetime.strptime(inicio, '%Y-%m-%d').date(),
                                datetime.strptime(fim, '%Y-%m-%d').date()))
        return ferias

    def get_folgas_ciclo(self):
        """Obtém os ciclos de folgas da base de dados usando semana_id (AAAASS)"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT p.nome, fc.semana_id, fc.dia_semana
            FROM folgas_ciclo fc
            JOIN pessoas p ON fc.pessoa_id = p.id
            WHERE p.ativo = 1
            ORDER BY p.nome, fc.semana_id, fc.dia_semana
        ''')
        folgas_data = cursor.fetchall()
        conn.close()

        ciclos = {}
        for nome, semana_id, dia_semana in folgas_data:
            if nome not in ciclos:
                ciclos[nome] = {}
            if semana_id not in ciclos[nome]:
                ciclos[nome][semana_id] = []
            ciclos[nome][semana_id].append(dia_semana)
        return ciclos

    def get_horarios_fixos(self):
        """Obtém os horários fixos da base de dados"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT p.nome, hf.data, hf.horario
            FROM horarios_fixos hf
            JOIN pessoas p ON hf.pessoa_id = p.id
        ''')
        horarios_data = cursor.fetchall()
        conn.close()
        horarios = {}
        for nome, data_str, horario in horarios_data:
            data = datetime.strptime(data_str, '%Y-%m-%d')
            horarios[(data, nome)] = horario
        return horarios

    def get_loja_fechada(self):
        """Obtém dias fechados com descrição"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT data, descricao FROM dias_loja_fechada ORDER BY data')
        dados = cursor.fetchall()
        conn.close()
        return [(datetime.strptime(d[0], '%Y-%m-%d'), d[1]) for d in dados]

class ScheduleWorker(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(object)
    error = pyqtSignal(str)

    def __init__(self, generator):
        super().__init__()
        self.generator = generator

    def run(self):
        try:
            self.generator.generate_schedule()
            df = self.generator.create_dataframe()
            self.finished.emit(df)
        except Exception as e:
            self.error.emit(str(e))

class WorkScheduleGenerator:
    def __init__(self):
        self.db = DatabaseManager()        
        self.start_date = datetime(2025, 11, 17)  # Segunda-feira
        self.num_semanas = 12
        self.pessoas = self.db.get_pessoas()
        self.ferias = self.db.get_ferias()
        self.horarios_fixos = self.db.get_horarios_fixos()
        self.loja_fechada_dates = self.db.get_loja_fechada()
        self.dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
        self.schedule_data = []

    def is_folga(self, pessoa, current_date):
        """Verifica se é folga baseado na base de dados (semana_id AAAASS)"""
        if pessoa not in self.ciclos_folgas:
            return False

        # Calcula semana_id no formato AAAASS
        year, iso_week, _ = current_date.isocalendar()
        semana_id = f"{year}{iso_week:02d}"

        dia_semana = current_date.weekday()  # 0=Segunda, ..., 6=Domingo

        return dia_semana in self.ciclos_folgas[pessoa].get(semana_id, [])

    def is_ferias(self, pessoa, data):
        if pessoa not in self.ferias:
            return False
        for inicio, fim in self.ferias[pessoa]:
            if inicio <= data.date() <= fim:
                return True
        return False

    def is_loja_fechada(self, data):
        return any(d[0].date() == data.date() for d in self.loja_fechada_dates)

    def get_turno_antonio_quinzenal(self, total_days, day):
        if total_days <= 3:
            return '11:00 - 20:00'
        semanas_completas = (total_days - 4) // 7
        periodo_A = semanas_completas % 2 == 0
        if day in [5, 6]:
            return None
        return '09:00 - 18:00' if periodo_A else '11:00 - 20:00'

    def get_susana_schedule(self, total_days, day, current_date, is_folga):
        if is_folga or self.is_ferias('Susana A.', current_date):
            return 'FOLGA' if is_folga else 'FÉRIAS'
        if 0 <= day <= 4:
            return '05:00 - 14:00'
        else:
            return 'FOLGA'

    def get_antonia_schedule(self, total_days, day, current_date, is_folga):
        if is_folga or self.is_ferias('Antónia F.', current_date):
            return 'FOLGA' if is_folga else 'FÉRIAS'
        if day == 1:
            return '06:00 - 15:00'
        else:
            return '07:00 - 16:00'

    def get_antonio_schedule(self, total_days, day, current_date, is_folga, ds):
        if (current_date, 'António C.') in self.horarios_fixos:
            return self.horarios_fixos[(current_date, 'António C.')]

        if is_folga or self.is_ferias('António C.', current_date):
            return 'FOLGA' if is_folga else 'FÉRIAS'

        # Filtra horários válidos de trabalho dos outros
        outros_horarios = [
            h for p, h in ds.items()
            if p != 'António C.' and isinstance(h, str) and h not in ['FOLGA', 'FÉRIAS', 'Loja Fechada']
        ]

        has_early = any(h.startswith(('05:', '06:', '07:', '08:')) for h in outros_horarios)  # 08: conta como cedo no fds
        has_late  = any(' - 20:00' in h for h in outros_horarios)

        # ===================================================================
        # FIM-DE-SEMANA (Sábado = 5, Domingo = 6)
        # ===================================================================
        if day in [5, 6]:
            # Prioridade máxima: garantir fecho às 20:00
            if not has_late:
                return '11:00 - 20:00'
            # Só depois, se já houver quem feche, verifica se precisa abrir cedo
            if not has_early:
                return '08:00 - 17:00'
            # Se já houver abertura cedo e fecho tarde → horário normal
            return '09:00 - 18:00'

        # ===================================================================
        # DIAS DE SEMANA (o teu código antigo pode ficar quase igual)
        # ===================================================================
        # (o resto do código que já tinhas para dias de semana)
        turno_quinzenal = self.get_turno_antonio_quinzenal(total_days, day)

        # Se no turno quinzenal dele for suposto fazer tarde, mas já há alguém tarde → faz normal
        if turno_quinzenal.startswith('11:') and has_late:
            turno_quinzenal = '09:00 - 18:00'

        # (o teu código de ajuste com o dia anterior pode ficar exatamente como está)

        return turno_quinzenal

        turno_quinzenal = self.get_turno_antonio_quinzenal(total_days, day)
        if turno_quinzenal.startswith('11:') and has_late:
            turno_quinzenal = '09:00 - 18:00'

        # Ajuste com dia anterior
        if total_days > 0:
            data_anterior = current_date - timedelta(days=1)
            horario_anterior = None
            for registro in self.schedule_data:
                if registro['Data_obj'] == data_anterior and 'António C.' in registro:
                    horario_anterior = registro['António C.']
                    break
            if horario_anterior and isinstance(horario_anterior, str) and horario_anterior not in ['FOLGA', 'FÉRIAS', 'Loja Fechada']:
                try:
                    hora_anterior = int(horario_anterior.split(':')[0])
                    hora_sugerida = int(turno_quinzenal.split(':')[0])
                    diferenca = abs(hora_anterior - hora_sugerida)
                    if diferenca > 3:
                        if hora_sugerida > hora_anterior:
                            hora_sugerida = hora_anterior + 3
                        else:
                            hora_sugerida = max(9, hora_anterior - 3)
                        return f"{hora_sugerida:02d}:00 - {hora_sugerida + 9:02d}:00"
                except:
                    pass
        return turno_quinzenal

    def get_magda_schedule(self, total_days, day, current_date, is_folga, ds):
        if (current_date, 'Magda G.') in self.horarios_fixos:
            return self.horarios_fixos[(current_date, 'Magda G.')]

        if is_folga or self.is_ferias('Magda G.', current_date):
            return 'FOLGA' if is_folga else 'FÉRIAS'

        outros_horarios = [
            h for p, h in ds.items()
            if p != 'Magda G.' and isinstance(h, str) and h not in ['FOLGA', 'FÉRIAS']
        ]

        has_late = any(h.startswith(('11:', '12:')) for h in outros_horarios)
        has_early = any(h.startswith(('05:', '06:', '07:')) for h in outros_horarios)

        if has_late:
            hora_sugerida = 9
        else:
            if not has_early:
                hora_sugerida = 7
            elif not has_late:
                hora_sugerida = 11
            else:
                hora_sugerida = 8 if day % 2 == 0 else 9

        # Ajuste com dia anterior
        if total_days > 0:
            data_anterior = current_date - timedelta(days=1)
            horario_anterior = None
            for registro in self.schedule_data:
                if registro['Data_obj'] == data_anterior and 'Magda G.' in registro:
                    horario_anterior = registro['Magda G.']
                    break
            if horario_anterior and isinstance(horario_anterior, str) and horario_anterior not in ['FOLGA', 'FÉRIAS', 'Loja Fechada']:
                try:
                    hora_anterior = int(horario_anterior.split(':')[0])
                    diferenca = abs(hora_anterior - hora_sugerida)
                    if diferenca > 3:
                        if hora_sugerida > hora_anterior:
                            hora_sugerida = hora_anterior + 3
                        else:
                            hora_sugerida = max(7, hora_anterior - 3)
                except:
                    pass

        hora_sugerida = max(7, min(11, int(hora_sugerida)))
        hora_saida = hora_sugerida + 9
        return f"{hora_sugerida:02d}:00 - {hora_saida:02d}:00"

    def get_eduardo_schedule(self, total_days, day, current_date, is_folga, ds):
        # 1. Horário fixo (BD)
        if (current_date, 'Eduardo S.') in self.horarios_fixos:
            return self.horarios_fixos[(current_date, 'Eduardo S.')]

        # 2. Folga/Férias
        if is_folga or self.is_ferias('Eduardo S.', current_date):
            return 'FOLGA' if is_folga else 'FÉRIAS'

        # 3. Terça-feira (dia 1)
        if day == 1:
            return '05:00 - 14:00'

        # 4. Verifica se alguém tem horário tarde (11:00 ou 12:00)
        tem_tarde = any(
            (current_date, p) in self.horarios_fixos and 
            isinstance(self.horarios_fixos[(current_date, p)], str) and
            self.horarios_fixos[(current_date, p)].startswith(('11:', '12:')) and
            self.horarios_fixos[(current_date, p)] not in ['FOLGA', 'FÉRIAS']
            or isinstance(ds.get(p), str) and ds[p].startswith(('11:', '12:')) and ds[p] not in ['FOLGA', 'FÉRIAS']
            for p in ['António C.', 'Magda G.']
        )

        # 5. DEFAULT: 06:00-15:00 ou 11:00-20:00
        return '06:00 - 15:00' if tem_tarde else '11:00 - 20:00'

    def needs_early_coverage(self, ds):
        """Verifica se precisa de cobertura matinal (05:00-07:00)"""
        return not any(
            isinstance(h, str) and h.startswith(('05:', '06:', '07:'))
            for p, h in ds.items()
            if p in self.pessoas and h not in ['FOLGA', 'FÉRIAS']
        )

    def has_coverage_until_20(self, ds):
        """Verifica se há cobertura até às 20:00"""
        return any(
            isinstance(h, str) and ' - 20:00' in h
            for p, h in ds.items()
            if p in self.pessoas and h not in ['FOLGA', 'FÉRIAS']
        )

    def has_early_shift_05_06(self, ds):
        """Verifica se há turno muito cedo (05:00 ou 06:00)"""
        return any(
            isinstance(h, str) and h.startswith(('05:', '06:'))
            for p, h in ds.items()
            if p in self.pessoas and h not in ['FOLGA', 'FÉRIAS']
        )

    def has_late_shift_12(self, ds):
        """Verifica se há turno às 12:00"""
        return any(
            isinstance(h, str) and h.startswith('12:')
            for p, h in ds.items()
            if p in self.pessoas and h not in ['FOLGA', 'FÉRIAS']
        )

    def generate_schedule(self):
        self.schedule_data = []
        ordem_processamento = ['Susana A.', 'Antónia F.', 'António C.', 'Magda G.', 'Eduardo S.']

        # RECARREGA TUDO DA BASE DE DADOS
        self.pessoas = self.db.get_pessoas()
        self.ferias = self.db.get_ferias()
        self.ciclos_folgas = self.db.get_folgas_ciclo()
        self.horarios_fixos = self.db.get_horarios_fixos()
        self.loja_fechada_dates = self.db.get_loja_fechada()

        for week in range(self.num_semanas):
            for day in range(7):
                total_days = (week * 7) + day
                current_date = self.start_date + timedelta(days=total_days)

                ds = {
                    'Semana': week + 1,
                    'Data': current_date.strftime('%d/%m/%Y'),
                    'Dia': self.dias_semana[day],
                    'Data_obj': current_date
                }

                # === 1. LOJA FECHADA (com descrição) ===
                if self.is_loja_fechada(current_date):
                    descricao = next(
                        (d[1] for d in self.loja_fechada_dates if d[0].date() == current_date.date()),
                        "Loja Fechada"
                    )
                    for pessoa in self.pessoas:
                        ds[pessoa] = descricao
                    self.schedule_data.append(ds)
                    continue
                # =======================================

                # === 2. MARCAR HORÁRIOS FIXOS PARA PROTEÇÃO ===
                horarios_fixos_hoje = set()
                for pessoa in self.pessoas:
                    if (current_date, pessoa) in self.horarios_fixos:
                        ds[pessoa] = self.horarios_fixos[(current_date, pessoa)]
                        horarios_fixos_hoje.add(pessoa)
                # ==============================================

                # === 3. PRIMEIRA PASSAGEM: GERAR HORÁRIOS BASE ===
                for pessoa in ordem_processamento:
                    if pessoa in horarios_fixos_hoje:
                        continue

                    folga = self.is_folga(pessoa, current_date)

                    if pessoa == 'Susana A.':
                        horario = self.get_susana_schedule(total_days, day, current_date, folga)
                    elif pessoa == 'Antónia F.':
                        horario = self.get_antonia_schedule(total_days, day, current_date, folga)
                    elif pessoa == 'António C.':
                        horario = self.get_antonio_schedule(total_days, day, current_date, folga, ds)
                    elif pessoa == 'Magda G.':
                        horario = self.get_magda_schedule(total_days, day, current_date, folga, ds)
                    elif pessoa == 'Eduardo S.':
                        horario = self.get_eduardo_schedule(total_days, day, current_date, folga, ds)
                    else:
                        horario = 'FOLGA' if folga else ('FÉRIAS' if self.is_ferias(pessoa, current_date) else '09:00 - 18:00')

                    ds[pessoa] = horario
                # ==================================================

                

                # === 5. SEGUNDA PASSAGEM: OTIMIZAR COBERTURA ATÉ 20:00 ===
                # Verifica se falta cobertura até 20:00
                if not self.has_coverage_until_20(ds):
                    # Conta quantas pessoas estão efetivamente a trabalhar (não FOLGA/FÉRIAS)
                    pessoas_trabalhando = sum(
                        1 for p, h in ds.items()
                        if p in self.pessoas and isinstance(h, str) and h not in ['FOLGA', 'FÉRIAS']
                    )
                    
                    # Só otimiza se houver 3+ pessoas a trabalhar
                    # (Se houver só 2, é porque alguém está de folga/férias, não compensa estender)
                    if pessoas_trabalhando >= 3:
                        # Lista de candidatos para horário tarde
                        candidatos = []
                        
                        for pessoa in ['António C.', 'Magda G.', 'Eduardo S.']:
                            if pessoa in horarios_fixos_hoje:
                                continue
                            
                            h = ds.get(pessoa, '')
                            if h not in ['FOLGA', 'FÉRIAS']:
                                candidatos.append(pessoa)
                        
                        # Tenta ajustar o melhor candidato
                        if candidatos:
                            # Verifica se já tem turnos muito cedo (05:00-06:00) → 07:00 não é essencial
                            tem_turno_muito_cedo = self.has_early_shift_05_06(ds)
                            
                            # Verifica se já tem turno às 12:00 → 11:00 seria redundante
                            tem_turno_12h = self.has_late_shift_12(ds)
                            
                            # Prioridade: Eduardo > Magda > António
                            for pessoa in ['Eduardo S.', 'Magda G.', 'António C.']:
                                if pessoa not in candidatos:
                                    continue
                                
                                h = ds.get(pessoa, '')
                                
                                # Eduardo: se não está às 05:00-14:00, pode ir para 11:00-20:00
                                if pessoa == 'Eduardo S.' and not h.startswith('05:'):
                                    # Só muda se não for terça (dia 1)
                                    if day != 1 and not tem_turno_12h:
                                        ds[pessoa] = '11:00 - 20:00'
                                        break
                                
                                # Magda: pode ir para 11:00-20:00 se não tiver turno 12h
                                elif pessoa == 'Magda G.' and not tem_turno_12h:
                                    # Se já tem turno 05:00-06:00, Magda pode ir tarde
                                    if tem_turno_muito_cedo:
                                        ds[pessoa] = '11:00 - 20:00'
                                        break
                                
                                # António: pode ir para 11:00-20:00 se for o turno quinzenal dele
                                elif pessoa == 'António C.' and not tem_turno_12h:
                                    # Verifica se é dia de semana (não sáb/dom)
                                    if day not in [5, 6]:
                                        turno_quinzenal = self.get_turno_antonio_quinzenal(total_days, day)
                                        if turno_quinzenal and turno_quinzenal.startswith('11:'):
                                            # Se já tem 05:00-06:00, António pode ir tarde
                                            if tem_turno_muito_cedo:
                                                ds[pessoa] = '11:00 - 20:00'
                                                break
                                            
                # === 4. COBERTURA MATINAL (05:00-07:00) ===
                if self.needs_early_coverage(ds):
                    for pessoa in reversed(ordem_processamento):  # Eduardo é o último
                        if pessoa in horarios_fixos_hoje:
                            continue  # NUNCA MEXER EM FIXO

                        h = ds.get(pessoa, '')
                        if h not in ['FOLGA', 'FÉRIAS'] and not h.startswith(('05:', '06:', '07:')):
                            if pessoa == 'Eduardo S.':
                                ds[pessoa] = '06:00 - 15:00'
                            else:
                                ds[pessoa] = '07:00 - 16:00'
                            break  # Só uma pessoa precisa
                # ============================================
                # =========================================================

                self.schedule_data.append(ds)

    def create_dataframe(self):
        df = pd.DataFrame(self.schedule_data)
        return df[['Semana', 'Data', 'Dia'] + list(self.pessoas.keys())]

    def export_to_excel(self, filename=None):
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'escala_trabalho_{timestamp}.xlsx'
        df = self.create_dataframe()
        wb = Workbook()
        ws = wb.active
        ws.title = "Escala de Trabalho"
        headers = df.columns.tolist()
        ws.append(headers)
        header_fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for row_idx, row_data in enumerate(df.values, 2):
            ws.append(row_data.tolist())
            current_date = datetime.strptime(row_data[1], '%d/%m/%Y')
            for col_idx, pessoa in enumerate(headers[3:], 4):
                horario = row_data[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx)
                if (current_date, pessoa) in self.horarios_fixos:
                    cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                    cell.font = Font(bold=True, color='CC0000')
                if horario == 'Loja Fechada':
                    cell.fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
                    cell.font = Font(bold=True, color='990000')
        thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        colors = {pessoa: info['cor'] for pessoa, info in self.pessoas.items()}
        folga_fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')
        weekend_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
        ferias_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
        for row in range(2, len(df) + 2):
            dia_semana = ws[f'C{row}'].value
            is_weekend = dia_semana in ['Sábado', 'Domingo']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_num)
                cell.border = thin
                cell.alignment = Alignment(horizontal='center')
                if header in colors:
                    if cell.value == 'FOLGA':
                        cell.fill = folga_fill
                        cell.font = Font(bold=True, color='0C5460')
                    elif cell.value == 'FÉRIAS':
                        cell.fill = ferias_fill
                        cell.font = Font(bold=True, color='000000')
                    elif cell.value != 'Loja Fechada' and not (datetime.strptime(ws[f'B{row}'].value, '%d/%m/%Y'), header) in self.horarios_fixos:
                        cell.fill = PatternFill(start_color=colors[header], end_color=colors[header], fill_type='solid')
                elif header == 'Dia' and is_weekend:
                    cell.fill = weekend_fill
        ws.append([])
        ws.append(['Gerado em:', datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        for col in ['D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
        wb.save(filename)
        return filename        

class ScheduleTableWidget(QTableWidget):
    def __init__(self):
        super().__init__()
        self.setAlternatingRowColors(True)
        self.horizontalHeader().setStretchLastSection(True)

    def display_data(self, df):
        self.setRowCount(len(df))
        self.setColumnCount(len(df.columns))
        self.setHorizontalHeaderLabels(df.columns.tolist())
        for row_idx, row_data in enumerate(df.values):
            for col_idx, cell_data in enumerate(row_data):
                item = QTableWidgetItem(str(cell_data))
                if col_idx >= 3:
                    if cell_data == 'FOLGA':
                        item.setBackground(QColor(209, 236, 241))
                        item.setForeground(QColor(12, 84, 96))
                    elif cell_data == 'FÉRIAS':
                        item.setBackground(QColor(255, 215, 0))
                        item.setForeground(QColor(0, 0, 0))
                    elif cell_data == 'Loja Fechada':
                        item.setBackground(QColor(255, 102, 102))
                        item.setForeground(QColor(153, 0, 0))
                    else:
                        pessoa = df.columns[col_idx]
                        cores = {
                            'Susana A.': QColor(232, 245, 232),
                            'António C.': QColor(255, 243, 205),
                            'Antónia F.': QColor(212, 237, 218),
                            'Magda G.': QColor(204, 229, 255),
                            'Eduardo S.': QColor(240, 230, 255)
                        }
                        if pessoa in cores:
                            item.setBackground(cores[pessoa])
                self.setItem(row_idx, col_idx, item)
        self.resizeColumnsToContents()
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.generator = WorkScheduleGenerator()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Gerador de Escalas")
        self.setGeometry(100, 100, 1400, 900)
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        control_layout = QHBoxLayout()
        self.generate_btn = QPushButton("Gerar Escala")
        self.generate_btn.clicked.connect(self.generate_schedule)
        self.generate_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 8px; }")
        self.export_btn = QPushButton("Exportar para Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setStyleSheet("QPushButton { background-color: #2196F3; color: white; font-weight: bold; padding: 8px; }")
        self.export_btn.setEnabled(False)
        self.close_btn = QPushButton("Fechar")
        self.close_btn.clicked.connect(self.hide)  # Apenas esconde, não fecha
        self.close_btn.setStyleSheet("QPushButton { background-color: #f44336; color: white; font-weight: bold; padding: 8px; }")
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        date_label = QLabel("Data de Início:")
        date_label.setStyleSheet("font-weight: bold;")
        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate(2025, 11, 17))
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setStyleSheet("padding: 5px;")
        semanas_label = QLabel("Semanas:")
        semanas_label.setStyleSheet("font-weight: bold;")
        self.semanas_spin = QSpinBox()
        self.semanas_spin.setRange(1, 52)
        self.semanas_spin.setValue(12)
        self.semanas_spin.setStyleSheet("padding: 5px;")
        control_layout.addWidget(date_label)
        control_layout.addWidget(self.date_edit)
        control_layout.addSpacing(10)
        control_layout.addWidget(semanas_label)
        control_layout.addWidget(self.semanas_spin)
        control_layout.addSpacing(20)
        control_layout.addWidget(self.generate_btn)
        control_layout.addWidget(self.export_btn)
        control_layout.addWidget(self.close_btn)
        control_layout.addWidget(self.progress_bar)
        control_layout.addStretch()
        layout.addLayout(control_layout)
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)
        self.table_tab = QWidget()
        self.table_layout = QVBoxLayout(self.table_tab)
        self.table_widget = ScheduleTableWidget()
        self.table_layout.addWidget(self.table_widget)
        self.summary_tab = QWidget()
        self.summary_layout = QVBoxLayout(self.summary_tab)
        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        self.summary_layout.addWidget(self.summary_text)
        self.tabs.addTab(self.table_tab, "Escala Completa")
        self.tabs.addTab(self.summary_tab, "Resumo e Estatísticas")
        self.statusBar().showMessage("Pronto para gerar escala")

    def generate_schedule(self):
        qdate = self.date_edit.date()
        start_date = datetime(qdate.year(), qdate.month(), qdate.day())
        self.generator.start_date = start_date
        self.generator.num_semanas = self.semanas_spin.value()
        self.generator.pessoas = self.generator.db.get_pessoas()
        self.generator.ferias = self.generator.db.get_ferias()
        self.generator.ciclos_folgas = self.generator.db.get_folgas_ciclo()
        self.generator.horarios_fixos = self.generator.db.get_horarios_fixos()
        self.generator.loja_fechada_dates = self.generator.db.get_loja_fechada()
        self.statusBar().showMessage("A gerar escala...")
        self.progress_bar.setVisible(True)
        self.generate_btn.setEnabled(False)
        self.worker = ScheduleWorker(self.generator)
        self.worker.finished.connect(self.on_generation_finished)
        self.worker.error.connect(self.on_generation_error)
        self.worker.start()

    def on_generation_finished(self, df):
        self.table_widget.display_data(df)
        self.export_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.generate_btn.setEnabled(True)
        self.update_summary()
        self.statusBar().showMessage(f"Escala gerada com sucesso! {len(df)} dias processados.")

    def on_generation_error(self, error_msg):
        self.progress_bar.setVisible(False)
        self.generate_btn.setEnabled(True)
        QMessageBox.critical(self, "Erro", f"Erro ao gerar escala:\n{error_msg}")
        self.statusBar().showMessage("Erro ao gerar escala")

    def update_summary(self):
        if not self.generator.schedule_data:
            return
        summary_text = "RESUMO DA ESCALA DE TRABALHO\n"
        summary_text += "=" * 50 + "\n\n"
        summary_text += f"Período: {self.generator.start_date.strftime('%d/%m/%Y')} a "
        end_date = self.generator.start_date + timedelta(days=len(self.generator.schedule_data)-1)
        summary_text += f"{end_date.strftime('%d/%m/%Y')}\n"
        summary_text += f"Total de dias: {len(self.generator.schedule_data)} dias\n\n"
        summary_text += "RESUMO DE DIAS TRABALHADOS:\n"
        summary_text += "-" * 40 + "\n"
        for pessoa in self.generator.pessoas:
            total = sum(1 for d in self.generator.schedule_data
                        if d.get(pessoa) not in ['FOLGA', 'FÉRIAS', 'Loja Fechada'])
            horas = total * self.generator.pessoas[pessoa]['horas']
            summary_text += f"{pessoa:<12}: {total:2d} dias | {horas:3d} horas\n"
        self.summary_text.setText(summary_text)

    def export_to_excel(self):
        try:
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Exportar para Excel",
                f"escala_trabalho_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            if filename:
                if not filename.endswith('.xlsx'):
                    filename += '.xlsx'
                saved_file = self.generator.export_to_excel(filename)
                QMessageBox.information(self, "Sucesso", f"Escala exportada para:\n{saved_file}")
                self.statusBar().showMessage(f"Ficheiro exportado: {saved_file}")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar:\n{str(e)}")

# Variável global para manter a janela
_janela_gerador = None

def mostrar_gerador():
    global _janela_gerador
    app = QApplication.instance()
    if app is None:
        print("ERRO: QApplication não está rodando. Inicie pelo main.py")
        return

    # Se já existe, apenas reativa
    if _janela_gerador is not None:
        _janela_gerador.show()
        _janela_gerador.raise_()
        _janela_gerador.activateWindow()
        return

    # Cria nova janela
    _janela_gerador = MainWindow()
    
    # Garante que ao fechar, não seja destruída
    _janela_gerador.setAttribute(Qt.WA_DeleteOnClose, False)
    
    # Ao fechar, apenas esconde
    def on_close():
        _janela_gerador.hide()
    
    _janela_gerador.closeEvent = lambda event: (on_close(), event.ignore())

    _janela_gerador.show()

    # Garante que a janela não seja destruída
    _janela_gerador.destroyed.connect(lambda: globals().update(_janela_gerador=None))