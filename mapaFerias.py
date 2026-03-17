#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Módulo para visualização do Mapa de Férias com filtros
"""

import sys
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QTableWidget, QTableWidgetItem, 
                             QPushButton, QLabel, QMessageBox, QHeaderView,
                             QComboBox, QLineEdit, QFormLayout, QGroupBox)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QColor
from datetime import datetime
import sqlite3

class DatabaseManager:
    def __init__(self, db_path='escala_trabalho.db'):
        self.db_path = db_path

    def get_ferias(self, pessoa_filtro=None, ano_filtro=None):
        """Obtém as férias da base de dados com filtros opcionais"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT p.nome, f.data_inicio, f.data_fim
            FROM ferias f
            JOIN pessoas p ON f.pessoa_id = p.id
            WHERE p.ativo = 1
        '''
        
        params = []
        
        # Aplicar filtro por pessoa
        if pessoa_filtro and pessoa_filtro != "Todas":
            query += " AND p.nome = ?"
            params.append(pessoa_filtro)
        
        # Aplicar filtro por ano
        if ano_filtro and ano_filtro != "Todos":
            query += " AND (strftime('%Y', f.data_inicio) = ? OR strftime('%Y', f.data_fim) = ?)"
            params.extend([ano_filtro, ano_filtro])
        
        query += " ORDER BY p.nome, f.data_inicio"
        
        cursor.execute(query, params)
        ferias_data = cursor.fetchall()
        conn.close()

        ferias = {}
        for nome, inicio, fim in ferias_data:
            if nome not in ferias:
                ferias[nome] = []
            ferias[nome].append((
                datetime.strptime(inicio, '%Y-%m-%d').date(),
                datetime.strptime(fim, '%Y-%m-%d').date()
            ))
        return ferias

    def get_pessoas(self):
        """Obtém a lista de pessoas"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT nome FROM pessoas WHERE ativo = 1 ORDER BY nome
        ''')
        pessoas = [row[0] for row in cursor.fetchall()]
        conn.close()
        return pessoas

    def get_anos_ferias(self):
        """Obtém lista de anos únicos com férias"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT DISTINCT strftime('%Y', data_inicio) as ano
            FROM ferias
            UNION
            SELECT DISTINCT strftime('%Y', data_fim) as ano
            FROM ferias
            ORDER BY ano DESC
        ''')
        anos = [row[0] for row in cursor.fetchall() if row[0]]
        conn.close()
        return anos

class MapaFeriasWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Mapa de Férias com Filtros")
        self.setGeometry(100, 100, 1000, 700)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        
        # ===== TÍTULO =====
        title = QLabel("🏖️ MAPA DE FÉRIAS")
        title.setFont(QFont("Arial", 16, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("padding: 20px; color: #FF9800;")
        layout.addWidget(title)
        
        # ===== FILTROS =====
        filtros_group = QGroupBox("Filtros")
        filtros_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #FF9800;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #FF9800;
            }
        """)
        
        filtros_layout = QHBoxLayout()
        
        # Filtro por Pessoa
        pessoa_label = QLabel("Pessoa:")
        pessoa_label.setFont(QFont("Arial", 10, QFont.Bold))
        
        self.combo_pessoa = QComboBox()
        self.combo_pessoa.setMinimumWidth(200)
        self.combo_pessoa.setFont(QFont("Arial", 10))
        
        # Filtro por Ano
        ano_label = QLabel("Ano:")
        ano_label.setFont(QFont("Arial", 10, QFont.Bold))
        
        self.combo_ano = QComboBox()
        self.combo_ano.setMinimumWidth(150)
        self.combo_ano.setFont(QFont("Arial", 10))
        
        # Botão Aplicar Filtros
        btn_aplicar = QPushButton("🔍 Aplicar Filtros")
        btn_aplicar.setFont(QFont("Arial", 10))
        btn_aplicar.clicked.connect(self.aplicar_filtros)
        btn_aplicar.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        
        # Botão Limpar Filtros
        btn_limpar = QPushButton("🗑️ Limpar Filtros")
        btn_limpar.setFont(QFont("Arial", 10))
        btn_limpar.clicked.connect(self.limpar_filtros)
        btn_limpar.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
        """)
        
        filtros_layout.addWidget(pessoa_label)
        filtros_layout.addWidget(self.combo_pessoa)
        filtros_layout.addSpacing(20)
        filtros_layout.addWidget(ano_label)
        filtros_layout.addWidget(self.combo_ano)
        filtros_layout.addSpacing(20)
        filtros_layout.addWidget(btn_aplicar)
        filtros_layout.addWidget(btn_limpar)
        filtros_layout.addStretch()
        
        filtros_group.setLayout(filtros_layout)
        layout.addWidget(filtros_group)
        
        # ===== RESUMO ESTATÍSTICO =====
        self.resumo_label = QLabel("")
        self.resumo_label.setFont(QFont("Arial", 11))
        self.resumo_label.setStyleSheet("padding: 10px; background-color: #E3F2FD; border-radius: 5px;")
        self.resumo_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.resumo_label)
        
        # ===== TABELA =====
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.table)
        
        # ===== BOTÕES INFERIORES =====
        buttons_layout = QHBoxLayout()
        
        btn_atualizar = QPushButton("🔄 Atualizar")
        btn_atualizar.setFont(QFont("Arial", 10))
        btn_atualizar.clicked.connect(self.carregar_dados)
        btn_atualizar.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        
        btn_exportar = QPushButton("📊 Exportar Resumo")
        btn_exportar.setFont(QFont("Arial", 10))
        btn_exportar.clicked.connect(self.exportar_resumo)
        btn_exportar.setStyleSheet("""
            QPushButton {
                background-color: #9C27B0;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #7B1FA2;
            }
        """)

        # No buttons_layout, antes do btn_fechar, adicione:
        btn_excel = QPushButton("📊 Exportar Excel Completo")
        btn_excel.setFont(QFont("Arial", 10))
        btn_excel.clicked.connect(self.exportar_excel_completo)
        btn_excel.setStyleSheet("""
            QPushButton {
                background-color: #27AE60;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        buttons_layout.addWidget(btn_excel)
        
        # No método init_ui(), alterar a conexão do botão Fechar:
        btn_fechar = QPushButton("✖️ Fechar")
        btn_fechar.setFont(QFont("Arial", 10))
        btn_fechar.clicked.connect(self.hide)  # ALTERADO: de close() para hide()
        btn_fechar.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """)
        
        buttons_layout.addWidget(btn_atualizar)
        buttons_layout.addWidget(btn_exportar)
        buttons_layout.addStretch()
        buttons_layout.addWidget(btn_fechar)
        layout.addLayout(buttons_layout)
        
        # ===== CARREGAR DADOS INICIAIS =====
        self.carregar_filtros()
        self.carregar_dados()

    def carregar_filtros(self):
        """Carrega as opções nos combobox de filtros"""
        try:
            # Carregar pessoas
            pessoas = self.db.get_pessoas()
            self.combo_pessoa.clear()
            self.combo_pessoa.addItem("Todas")
            self.combo_pessoa.addItems(pessoas)
            
            # Carregar anos
            anos = self.db.get_anos_ferias()
            self.combo_ano.clear()
            self.combo_ano.addItem("Todos")
            self.combo_ano.addItems(anos)
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar filtros:\n{str(e)}")

    def carregar_dados(self):
        """Carrega as férias na tabela com base nos filtros atuais"""
        try:
            # Obter filtros selecionados
            pessoa_filtro = self.combo_pessoa.currentText()
            ano_filtro = self.combo_ano.currentText()
            
            if ano_filtro == "Todos":
                ano_filtro = None
            if pessoa_filtro == "Todas":
                pessoa_filtro = None
            
            # Obter dados filtrados
            ferias = self.db.get_ferias(pessoa_filtro, ano_filtro)
            pessoas = self.db.get_pessoas()
            
            if not ferias:
                self.table.setRowCount(1)
                self.table.setColumnCount(1)
                self.table.setHorizontalHeaderLabels(["Informação"])
                item = QTableWidgetItem("Não há férias marcadas para os filtros selecionados")
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(0, 0, item)
                self.atualizar_resumo([], pessoa_filtro, ano_filtro)
                return
            
            # Preparar dados para tabela
            rows = []
            total_dias = 0
            pessoas_com_ferias = []
            
            # Se filtro por pessoa específica, usar apenas essa
            pessoas_para_processar = [pessoa_filtro] if pessoa_filtro else pessoas
            
            for pessoa in pessoas_para_processar:
                if pessoa in ferias:
                    pessoas_com_ferias.append(pessoa)
                    for inicio, fim in ferias[pessoa]:
                        dias = (fim - inicio).days + 1
                        total_dias += dias
                        
                        # Verificar se passa pelo ano do filtro
                        if ano_filtro:
                            ano_inicio = inicio.year
                            ano_fim = fim.year
                            ano_filtro_int = int(ano_filtro)
                            
                            # Verificar se o período inclui o ano filtrado
                            if ano_inicio <= ano_filtro_int <= ano_fim:
                                rows.append({
                                    'Pessoa': pessoa,
                                    'Início': inicio.strftime('%d/%m/%Y'),
                                    'Fim': fim.strftime('%d/%m/%Y'),
                                    'Dias': dias
                                })
                        else:
                            rows.append({
                                'Pessoa': pessoa,
                                'Início': inicio.strftime('%d/%m/%Y'),
                                'Fim': fim.strftime('%d/%m/%Y'),
                                'Dias': dias
                            })
            
            # Configurar tabela
            self.table.setRowCount(len(rows))
            self.table.setColumnCount(5)
            self.table.setHorizontalHeaderLabels(['#', 'Pessoa', 'Início', 'Fim', 'Dias'])
            
            # Preencher tabela
            cores = {
                'Susana A.': QColor(232, 245, 232),
                'António C.': QColor(255, 243, 205),
                'Antónia F.': QColor(212, 237, 218),
                'Magda G.': QColor(204, 229, 255),
                'Eduardo S.': QColor(240, 230, 255)
            }
            
            for row_idx, row_data in enumerate(rows):
                # Número da linha
                item_num = QTableWidgetItem(str(row_idx + 1))
                item_num.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, 0, item_num)
                
                # Pessoa
                item_pessoa = QTableWidgetItem(row_data['Pessoa'])
                if row_data['Pessoa'] in cores:
                    item_pessoa.setBackground(cores[row_data['Pessoa']])
                self.table.setItem(row_idx, 1, item_pessoa)
                
                # Início
                item_inicio = QTableWidgetItem(row_data['Início'])
                item_inicio.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, 2, item_inicio)
                
                # Fim
                item_fim = QTableWidgetItem(row_data['Fim'])
                item_fim.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, 3, item_fim)
                
                # Dias
                item_dias = QTableWidgetItem(str(row_data['Dias']))
                item_dias.setTextAlignment(Qt.AlignCenter)
                item_dias.setFont(QFont("Arial", 10, QFont.Bold))
                self.table.setItem(row_idx, 4, item_dias)
            
            # Ajustar colunas
            self.table.resizeColumnsToContents()
            self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            
            # Atualizar resumo
            self.atualizar_resumo(rows, pessoa_filtro, ano_filtro, total_dias)
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar dados:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def atualizar_resumo(self, rows, pessoa_filtro, ano_filtro, total_dias=0):
        """Atualiza o label com o resumo estatístico"""
        try:
            num_registros = len(rows)
            
            # Construir texto do resumo
            resumo_parts = []
            resumo_parts.append(f"📊 <b>RESUMO:</b>")
            
            if pessoa_filtro:
                resumo_parts.append(f"<b>Pessoa:</b> {pessoa_filtro}")
            else:
                resumo_parts.append("<b>Pessoa:</b> Todas")
            
            if ano_filtro:
                resumo_parts.append(f"<b>Ano:</b> {ano_filtro}")
            else:
                resumo_parts.append("<b>Ano:</b> Todos")
            
            resumo_parts.append(f"<b>Períodos:</b> {num_registros}")
            resumo_parts.append(f"<b>Total Dias:</b> {total_dias}")
            
            # Calcular média se houver registros
            if num_registros > 0:
                media_dias = total_dias / num_registros
                resumo_parts.append(f"<b>Média por Período:</b> {media_dias:.1f} dias")
            
            self.resumo_label.setText(" | ".join(resumo_parts))
            
        except Exception as e:
            self.resumo_label.setText(f"Erro ao calcular resumo: {str(e)}")

    def aplicar_filtros(self):
        """Aplica os filtros selecionados"""
        self.carregar_dados()

    def limpar_filtros(self):
        """Limpa todos os filtros"""
        self.combo_pessoa.setCurrentIndex(0)  # "Todas"
        self.combo_ano.setCurrentIndex(0)     # "Todos"
        self.carregar_dados()

    def exportar_resumo(self):
        """Exporta o resumo para um ficheiro de texto"""
        try:
            from PyQt5.QtWidgets import QFileDialog
            
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Exportar Resumo",
                f"resumo_ferias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                "Text Files (*.txt)"
            )
            
            if filename:
                if not filename.endswith('.txt'):
                    filename += '.txt'
                
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write("=" * 50 + "\n")
                    f.write("RESUMO DE FÉRIAS\n")
                    f.write("=" * 50 + "\n\n")
                    
                    # Filtros aplicados
                    pessoa = self.combo_pessoa.currentText()
                    ano = self.combo_ano.currentText()
                    f.write(f"FILTROS APLICADOS:\n")
                    f.write(f"  Pessoa: {pessoa}\n")
                    f.write(f"  Ano: {ano}\n\n")
                    
                    # Estatísticas
                    total_linhas = self.table.rowCount()
                    total_dias = 0
                    
                    f.write("DETALHE DAS FÉRIAS:\n")
                    f.write("-" * 50 + "\n")
                    
                    for row in range(total_linhas):
                        try:
                            pessoa_item = self.table.item(row, 1)
                            inicio_item = self.table.item(row, 2)
                            fim_item = self.table.item(row, 3)
                            dias_item = self.table.item(row, 4)
                            
                            if all([pessoa_item, inicio_item, fim_item, dias_item]):
                                pessoa = pessoa_item.text()
                                inicio = inicio_item.text()
                                fim = fim_item.text()
                                dias = int(dias_item.text())
                                total_dias += dias
                                
                                f.write(f"{row+1:3d}. {pessoa:<15} {inicio} - {fim} ({dias:3d} dias)\n")
                        except:
                            continue
                    
                    f.write("\n" + "=" * 50 + "\n")
                    f.write(f"TOTAL DE PERÍODOS: {total_linhas}\n")
                    f.write(f"TOTAL DE DIAS: {total_dias}\n")
                    
                    if total_linhas > 0:
                        f.write(f"MÉDIA POR PERÍODO: {total_dias/total_linhas:.1f} dias\n")
                    
                    f.write("=" * 50 + "\n")
                    f.write(f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                
                QMessageBox.information(self, "Sucesso", f"Resumo exportado para:\n{filename}")
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar:\n{str(e)}")


    def exportar_excel_completo(self):
        """Exporta o mapa de férias completo para Excel com todos os colaboradores"""
        try:
            from PyQt5.QtWidgets import QFileDialog
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.styles import numbers
            
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Exportar Mapa de Férias para Excel",
                f"mapa_ferias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not filename:
                return
            
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            # Obter dados completos (sem filtros)
            ferias = self.db.get_ferias()
            pessoas = self.db.get_pessoas()
            
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Mapa de Férias"
            
            # Cabeçalhos
            headers = ['Pessoa', 'Início', 'Fim', 'Dias', 'Período']
            ws.append(headers)
            
            # Formatar cabeçalhos
            header_fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Cores para cada pessoa
            cores = {
                'Susana A.': 'E8F5E8',
                'António C.': 'FFF3CD',
                'Antónia F.': 'D4EDDA',
                'Magda G.': 'CCE5FF',
                'Eduardo S.': 'F0E6FF'
            }
            
            # Adicionar dados
            for pessoa in pessoas:
                if pessoa in ferias:
                    for inicio, fim in ferias[pessoa]:
                        dias = (fim - inicio).days + 1
                        periodo = f"{inicio.strftime('%d/%m')} a {fim.strftime('%d/%m')}"
                        
                        row_data = [
                            pessoa,
                            inicio,  # Passar objeto date em vez de string
                            fim,     # Passar objeto date em vez de string
                            dias,
                            periodo
                        ]
                        ws.append(row_data)
            
            # Formatar células
            thin = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(2, ws.max_row + 1):
                pessoa = ws[f'A{row}'].value
                
                # Formatar coluna B (Início) como data
                cell_b = ws[f'B{row}']
                cell_b.number_format = 'DD-MM-YYYY'  # Formato para LibreOffice/Excel
                cell_b.border = thin
                cell_b.alignment = Alignment(horizontal='center')
                
                # Formatar coluna C (Fim) como data
                cell_c = ws[f'C{row}']
                cell_c.number_format = 'DD-MM-YYYY'  # Formato para LibreOffice/Excel
                cell_c.border = thin
                cell_c.alignment = Alignment(horizontal='center')
                
                # Formatar outras colunas
                for col in [1, 4, 5]:
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Aplicar cor por pessoa na coluna A
                    if col == 1 and pessoa in cores:
                        cell.fill = PatternFill(start_color=cores[pessoa], end_color=cores[pessoa], fill_type='solid')
                    
                    # Negrito para dias
                    if col == 4:
                        cell.font = Font(bold=True)
            
            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 18  # Pessoa
            ws.column_dimensions['B'].width = 12  # Início
            ws.column_dimensions['C'].width = 12  # Fim
            ws.column_dimensions['D'].width = 8   # Dias
            ws.column_dimensions['E'].width = 20  # Período
            
            # === SEGUNDA ABA: VISÃO POR PESSOA (lado a lado) ===
            ws2 = wb.create_sheet("Visão por Pessoa")
            
            # Cabeçalhos
            headers2 = ['Período'] + pessoas
            ws2.append(headers2)
            
            # Formatar cabeçalhos
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Coletar todos os períodos únicos com data de início para ordenação
            todos_periodos = []
            periodos_por_pessoa = {}
            
            for pessoa in pessoas:
                periodos_por_pessoa[pessoa] = []
                if pessoa in ferias:
                    for inicio, fim in ferias[pessoa]:
                        dias = (fim - inicio).days + 1
                        periodo = f"{inicio.strftime('%d/%m')} a {fim.strftime('%d/%m')}"
                        periodos_por_pessoa[pessoa].append({
                            'periodo': periodo,
                            'dias': dias,
                            'data_inicio': inicio
                        })
                        # Guardar com data para ordenação
                        todos_periodos.append({
                            'periodo': periodo,
                            'data_inicio': inicio
                        })
            
            # Remover duplicados e ordenar por data
            periodos_unicos = []
            periodos_set = set()
            for p in sorted(todos_periodos, key=lambda x: x['data_inicio']):
                if p['periodo'] not in periodos_set:
                    periodos_unicos.append(p['periodo'])
                    periodos_set.add(p['periodo'])
            
            # Preencher dados
            for periodo in periodos_unicos:
                row_data = [periodo]
                for pessoa in pessoas:
                    encontrado = False
                    for p in periodos_por_pessoa[pessoa]:
                        if p['periodo'] == periodo:
                            row_data.append(f"✓ ({p['dias']}d)")
                            encontrado = True
                            break
                    if not encontrado:
                        row_data.append("-")
                ws2.append(row_data)
            
            # Formatar segunda aba
            for row in range(2, ws2.max_row + 1):
                for col in range(1, len(pessoas) + 2):
                    cell = ws2.cell(row=row, column=col)
                    cell.border = thin
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Aplicar cores por pessoa
                    if col > 1:
                        pessoa = pessoas[col-2]
                        if cell.value != "-" and pessoa in cores:
                            cell.fill = PatternFill(start_color=cores[pessoa], end_color=cores[pessoa], fill_type='solid')
                            cell.font = Font(bold=True)
            
            # Ajustar larguras
            ws2.column_dimensions['A'].width = 20
            for i, pessoa in enumerate(pessoas, 2):
                ws2.column_dimensions[chr(64 + i)].width = 15
            
            # === TERCEIRA ABA: RESUMO ===
            ws3 = wb.create_sheet("Resumo")
            
            ws3.append(['RESUMO DE FÉRIAS'])
            ws3.append([])
            ws3.append(['Pessoa', 'Total de Dias', 'Nº de Períodos', 'Média por Período'])
            
            total_geral_dias = 0
            total_geral_periodos = 0
            
            for pessoa in pessoas:
                if pessoa in ferias:
                    total_dias = 0
                    num_periodos = len(ferias[pessoa])
                    for inicio, fim in ferias[pessoa]:
                        total_dias += (fim - inicio).days + 1
                    
                    media = total_dias / num_periodos if num_periodos > 0 else 0
                    
                    ws3.append([pessoa, total_dias, num_periodos, f"{media:.1f}"])
                    
                    total_geral_dias += total_dias
                    total_geral_periodos += num_periodos
                else:
                    ws3.append([pessoa, 0, 0, "0"])
            
            ws3.append([])
            ws3.append(['TOTAL GERAL', total_geral_dias, total_geral_periodos, f"{total_geral_dias/total_geral_periodos:.1f}" if total_geral_periodos > 0 else "0"])
            
            # Formatar resumo
            for row in range(1, ws3.max_row + 1):
                for col in range(1, 5):
                    cell = ws3.cell(row=row, column=col)
                    cell.border = thin
                    cell.alignment = Alignment(horizontal='center')
            
            # Título em negrito
            ws3['A1'].font = Font(bold=True, size=14)
            ws3.merge_cells('A1:D1')
            
            # Formatar totais
            last_row = ws3.max_row
            for col in range(1, 5):
                cell = ws3.cell(row=last_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            
            ws3.column_dimensions['A'].width = 18
            ws3.column_dimensions['B'].width = 12
            ws3.column_dimensions['C'].width = 15
            ws3.column_dimensions['D'].width = 15
            
            # Salvar
            wb.save(filename)
            
            QMessageBox.information(
                self, 
                "Sucesso", 
                f"Mapa de férias exportado com sucesso!\n\n"
                f"Arquivo: {filename}\n\n"
                f"O ficheiro contém 3 abas:\n"
                f"• Mapa de Férias - Lista detalhada\n"
                f"• Visão por Pessoa - Períodos lado a lado\n"
                f"• Resumo - Estatísticas por pessoa\n\n"
                f"Nota: As datas estão formatadas como DD-MM-YYYY"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar Excel:\n{str(e)}")
            import traceback
            traceback.print_exc()

# Variável global para manter a janela
_janela_mapa_ferias = None

def mostrar_mapa_ferias():
    """Função para mostrar a janela do Mapa de Férias"""
    global _janela_mapa_ferias
    
    # Se já existe uma janela, apenas reativa
    if _janela_mapa_ferias is not None:
        _janela_mapa_ferias.show()
        _janela_mapa_ferias.raise_()
        _janela_mapa_ferias.activateWindow()
        return
    
    # Cria nova janela
    _janela_mapa_ferias = MapaFeriasWindow()
    
    # Garante que ao fechar, não seja destruída
    _janela_mapa_ferias.setAttribute(Qt.WA_DeleteOnClose, False)
    
    # Ao fechar, apenas esconde
    def on_close():
        _janela_mapa_ferias.hide()
    
    _janela_mapa_ferias.closeEvent = lambda event: (on_close(), event.ignore())
    
    _janela_mapa_ferias.show()

    # Limpa a referência quando a janela for realmente destruída
    _janela_mapa_ferias.destroyed.connect(lambda: globals().update(_janela_mapa_ferias=None))

if __name__ == "__main__":
    # Para testar individualmente
    app = QApplication(sys.argv)
    window = MapaFeriasWindow()
    window.show()
    sys.exit(app.exec_())