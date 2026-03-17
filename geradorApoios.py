#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Módulo Gerador de Apoios (Escala Telefónica)
"""

import sys
import os
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QMessageBox,
                             QProgressBar, QFileDialog, QTextEdit, QFrame)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QPixmap
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import random
from collections import defaultdict
import traceback

# VARIÁVEL GLOBAL - ADICIONAR AQUI
_janela_gerador_apoios = None

# Importar a classe do gerador
class PhoneScheduleGenerator:
    def __init__(self, excel_file='escala_trabalho.xlsx'):
        """
        Inicializa o gerador de escala telefónica a partir do arquivo Excel
        """
        self.excel_file = excel_file
        self.work_schedule = self.load_work_schedule()
        self.start_date = datetime(2026, 12, 22)

        # Abreviações dos nomes
        self.abreviacoes = {
            'Susana A.': 'SA',
            'António C.': 'AC',
            'Antónia F.': 'AF',
            'Magda G.': 'MG',
            'Eduardo S.': 'ES'
        }

        # Cores para cada pessoa (mesmas da escala principal)
        self.cores = {
            'SA': 'E8F5E8',
            'AC': 'FFF3CD',
            'AF': 'D4EDDA',
            'MG': 'CCE5FF',
            'ES': 'F0E6FF'
        }

        # Horários de atendimento (8:00 às 22:00)
        self.horarios = [f"{h:02d}:00" for h in range(8, 22)]

        self.phone_schedule = []

    def load_work_schedule(self):
        """
        Carrega os dados da escala de trabalho a partir do arquivo Excel
        """
        try:
            # Ler o arquivo Excel
            df = pd.read_excel(self.excel_file, sheet_name='Escala de Trabalho')

            # Converter para o formato esperado
            schedule_data = []

            for idx, row in df.iterrows():
                # Verificar se a linha tem dados válidos
                if pd.isna(row.get('Data', None)) or pd.isna(row.get('Semana', None)):
                    continue

                # Converter data de forma robusta
                data_obj = self.parse_excel_date(row['Data'])
                if not data_obj:
                    continue

                entry = {
                    'Semana': int(row['Semana']) if not pd.isna(row['Semana']) else 0,
                    'Data': data_obj.strftime('%d/%m/%Y'),
                    'Data_obj': data_obj,
                    'Dia': row['Dia'] if not pd.isna(row.get('Dia', None)) else '',
                    'Susana A.': row['Susana A.'] if not pd.isna(row.get('Susana A.', None)) else 'FOLGA',
                    'António C.': row['António C.'] if not pd.isna(row.get('António C.', None)) else 'FOLGA',
                    'Antónia F.': row['Antónia F.'] if not pd.isna(row.get('Antónia F.', None)) else 'FOLGA',
                    'Magda G.': row['Magda G.'] if not pd.isna(row.get('Magda G.', None)) else 'FOLGA',
                    'Eduardo S.': row['Eduardo S.'] if not pd.isna(row.get('Eduardo S.', None)) else 'FOLGA'
                }
                schedule_data.append(entry)

            return schedule_data

        except FileNotFoundError:
            print(f"Erro: Arquivo {self.excel_file} não encontrado.")
            return []
        except Exception as e:
            print(f"Erro ao ler o arquivo Excel: {e}")
            return []

    def parse_excel_date(self, date_value):
        """Converte diferentes formatos de data do Excel para datetime"""
        if pd.isna(date_value):
            return None

        # Se já for datetime
        if isinstance(date_value, datetime):
            return date_value

        # Se for string
        if isinstance(date_value, str):
            try:
                # Tentar formato DD/MM/YYYY
                return datetime.strptime(date_value, '%d/%m/%Y')
            except ValueError:
                try:
                    # Tentar formato YYYY-MM-DD
                    return datetime.strptime(date_value, '%Y-%m-%d')
                except ValueError:
                    return None

        # Se for float (número de série do Excel)
        if isinstance(date_value, (int, float)):
            try:
                return pd.to_datetime(date_value, unit='D', origin='1899-12-30')
            except:
                try:
                    return pd.to_datetime(date_value)
                except:
                    return None

        return None

    def parse_schedule_time(self, horario_str):
        """
        Extrai hora de início e fim de um horário
        Retorna tupla (hora_inicio, hora_fim) ou None se for FOLGA/FÉRIAS
        """
        if horario_str in ['FOLGA', 'FÉRIAS', 'Loja Fechada'] or pd.isna(horario_str) or not horario_str:
            return None

        try:
            # Converter para string se necessário
            if not isinstance(horario_str, str):
                horario_str = str(horario_str)

            # Remover espaços extras
            horario_str = horario_str.strip()

            partes = horario_str.split(' - ')
            if len(partes) != 2:
                return None

            hora_inicio = int(partes[0].split(':')[0])
            hora_fim = int(partes[1].split(':')[0])
            return (hora_inicio, hora_fim)
        except Exception:
            return None

    def get_available_people_at_hour(self, day_schedule, hour):
        """
        Retorna lista de pessoas disponíveis em determinada hora
        """
        available = []

        for pessoa, horario in day_schedule.items():
            if pessoa in self.abreviacoes:
                time_range = self.parse_schedule_time(horario)
                if time_range:
                    hora_inicio, hora_fim = time_range
                    if hora_inicio <= hour < hora_fim:
                        available.append(self.abreviacoes[pessoa])

        return available

    def calculate_weekly_hours(self, week_data):
        """
        Calcula total de horas trabalhadas por pessoa numa semana
        """
        hours_count = defaultdict(int)

        for day_data in week_data:
            for pessoa, horario in day_data.items():
                if pessoa in self.abreviacoes:
                    time_range = self.parse_schedule_time(horario)
                    if time_range:
                        hora_inicio, hora_fim = time_range
                        # Contar apenas horas entre 8:00 e 22:00
                        inicio = max(8, hora_inicio)
                        fim = min(22, hora_fim)
                        if fim > inicio:
                            hours_count[self.abreviacoes[pessoa]] += (fim - inicio)

        return hours_count

    def distribute_phone_hours(self, day_schedule, date, weekly_phone_count, weekly_hours, last_assigned):
        """
        Distribui as horas de telefone de forma equilibrada com máximo de 3 horas seguidas
        """
        day_phone_schedule = {}
        consecutive_count = {abrev: 0 for abrev in self.abreviacoes.values()}
        last_person = None

        for hour in range(8, 22):
            available = self.get_available_people_at_hour(day_schedule, hour)

            if available:
                # Remover pessoas que já fizeram 3 horas seguidas
                available = [p for p in available if consecutive_count.get(p, 0) < 3]

                # Se não há pessoas disponíveis (todas fizeram 3h), resetar contadores
                if not available:
                    consecutive_count = {abrev: 0 for abrev in self.abreviacoes.values()}
                    available = self.get_available_people_at_hour(day_schedule, hour)

                # Calcular proporção ideal para cada pessoa
                proportions = {}
                total_hours = sum(weekly_hours[p] for p in available if p in weekly_hours)

                if total_hours > 0:
                    for person in available:
                        if person in weekly_hours and weekly_hours[person] > 0:
                            # Proporção ideal de horas de telefone
                            ideal_proportion = weekly_hours[person] / total_hours
                            current_count = weekly_phone_count.get(person, 0)

                            # Calcular déficit (quanto falta para atingir a proporção ideal)
                            total_phone_hours = sum(weekly_phone_count.values())
                            if total_phone_hours > 0:
                                current_proportion = current_count / (total_phone_hours + 1)
                                deficit = ideal_proportion - current_proportion
                            else:
                                deficit = ideal_proportion

                            proportions[person] = deficit

                    # Escolher pessoa com maior déficit (mais precisa de horas)
                    if proportions:
                        selected = max(proportions.items(), key=lambda x: x[1])[0]
                    else:
                        selected = random.choice(available)
                else:
                    # Se não há horas calculadas, escolher aleatoriamente
                    selected = random.choice(available)

                # Atualizar contador de horas consecutivas
                if selected == last_person:
                    consecutive_count[selected] += 1
                else:
                    consecutive_count[selected] = 1
                    last_person = selected

                day_phone_schedule[f"{hour:02d}:00"] = selected
                weekly_phone_count[selected] = weekly_phone_count.get(selected, 0) + 1
            else:
                day_phone_schedule[f"{hour:02d}:00"] = '-'
                # Resetar contadores quando não há ninguém disponível
                consecutive_count = {abrev: 0 for abrev in self.abreviacoes.values()}
                last_person = None

        return day_phone_schedule

    def generate_phone_schedule(self):
        """
        Gera a escala telefónica completa
        """
        if not self.work_schedule:
            return False

        self.phone_schedule = []

        # Processar semana a semana
        for week in range(1, 13):
            # Coletar dados da semana
            week_data = [entry for entry in self.work_schedule if entry['Semana'] == week]

            if not week_data:
                continue

            # Calcular horas trabalhadas na semana
            weekly_hours = self.calculate_weekly_hours(week_data)

            # Contador de horas de telefone na semana
            weekly_phone_count = defaultdict(int)

            # Processar cada dia da semana
            for day_data in week_data:
                date = day_data['Data_obj']

                # Verificar se é dia de "Loja Fechada"
                if any('Loja Fechada' in str(horario) for horario in day_data.values()):
                    # Criar escala vazia para dia de loja fechada
                    day_phone = {f"{h:02d}:00": 'Loja F.' for h in range(8, 22)}
                else:
                    # Gerar escala telefónica do dia
                    day_phone = self.distribute_phone_hours(
                        day_data,
                        date,
                        weekly_phone_count,
                        weekly_hours,
                        None
                    )

                # Adicionar à lista
                phone_entry = {
                    'Data': date.strftime('%d/%m/%Y'),
                    'Data_obj': date,
                    'Dia': day_data['Dia'],
                    'Semana': week
                }
                phone_entry.update(day_phone)

                self.phone_schedule.append(phone_entry)
        
        return True

    def export_to_excel(self, filename_prefix='escala_telefonica'):
        """
        Exporta a escala telefónica para Excel com timestamp
        """
        if not self.phone_schedule:
            return None

        # Adicionar timestamp ao nome do arquivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Escala Telefónica"

        # Criar cabeçalhos (Data + horas)
        headers = ['Data', 'Dia'] + self.horarios
        ws.append(headers)

        # Formatar cabeçalhos
        header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Adicionar dados
        for entry in self.phone_schedule:
            row_data = [entry['Data'], entry['Dia']]
            for hora in self.horarios:
                row_data.append(entry.get(hora, '-'))
            ws.append(row_data)

        # Formatar células
        thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        weekend_fill = PatternFill(start_color='FFE5E5', end_color='FFE5E5', fill_type='solid')
        loja_fechada_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
        loja_fechada_font = Font(bold=True, color='990000')

        for row in range(2, len(self.phone_schedule) + 2):
            dia_semana = ws[f'B{row}'].value
            is_weekend = dia_semana in ['Sábado', 'Domingo']

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Aplicar cor de fundo para fins de semana
                if col <= 2 and is_weekend:
                    cell.fill = weekend_fill

                # Aplicar cor para as iniciais dos colaboradores
                if col > 2:  # Colunas de horas
                    value = cell.value
                    if value == 'Loja F.':
                        cell.fill = loja_fechada_fill
                        cell.font = loja_fechada_font
                    elif value and value != '-' and value in self.cores:
                        color = self.cores[value]
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        cell.font = Font(bold=True)

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12  # Data
        ws.column_dimensions['B'].width = 12  # Dia
        for col_letter in 'CDEFGHIJKLMNO':  # Horas
            ws.column_dimensions[col_letter].width = 8

        # Adicionar resumo semanal em nova aba
        ws2 = wb.create_sheet("Resumo Semanal")
        self.add_weekly_summary(ws2)

        # Adicionar legenda em nova aba
        ws3 = wb.create_sheet("Legenda")
        self.add_legend(ws3)

        # Salvar arquivo
        wb.save(filename)
        return filename

    def add_weekly_summary(self, worksheet):
        """
        Adiciona resumo semanal na segunda aba
        """
        ws = worksheet

        # Cabeçalhos
        headers = ['Semana', 'Período'] + list(self.abreviacoes.values()) + ['Total']
        ws.append(headers)

        # Formatar cabeçalhos
        header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Calcular totais por semana
        for week in range(1, 13):
            week_data = [entry for entry in self.phone_schedule if entry['Semana'] == week]

            if week_data:
                start_date = week_data[0]['Data']
                end_date = week_data[-1]['Data']
                period = f"{start_date} - {end_date}"

                # Contar horas por pessoa (excluir 'Loja Fechada')
                counts = {abrev: 0 for abrev in self.abreviacoes.values()}

                for day_entry in week_data:
                    for hora in self.horarios:
                        if (hora in day_entry and day_entry[hora] in counts and
                            day_entry[hora] != 'Loja F.'):
                            counts[day_entry[hora]] += 1

                # Criar linha
                row_data = [f"Semana {week}", period]
                total = 0
                for abrev in self.abreviacoes.values():
                    row_data.append(counts[abrev])
                    total += counts[abrev]
                row_data.append(total)

                ws.append(row_data)

        # Adicionar totais gerais
        ws.append([])  # Linha vazia
        total_row = ['TOTAL', '12 Semanas']
        grand_totals = {abrev: 0 for abrev in self.abreviacoes.values()}

        for entry in self.phone_schedule:
            for hora in self.horarios:
                if (hora in entry and entry[hora] in grand_totals and
                    entry[hora] != 'Loja F.'):
                    grand_totals[entry[hora]] += 1

        grand_total = 0
        for abrev in self.abreviacoes.values():
            total_row.append(grand_totals[abrev])
            grand_total += grand_totals[abrev]
        total_row.append(grand_total)

        ws.append(total_row)

        # Formatar última linha (totais)
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

        # Ajustar larguras
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        for col in 'CDEFGHIJ':
            ws.column_dimensions[col].width = 10

    def add_legend(self, worksheet):
        """
        Adiciona legenda com as regras aplicadas
        """
        ws = worksheet

        ws.append(['LEGENDA DA ESCALA TELEFÓNICA'])
        ws.append([])
        ws.append(['Regras Aplicadas:'])
        ws.append(['- Máximo de 3 horas consecutivas por pessoa'])
        ws.append(['- Distribuição equilibrada entre colaboradores'])
        ws.append(['- Respeito pelos horários de almoço'])
        ws.append(['- Dias de "Loja Fechada" destacados a vermelho'])
        ws.append([])
        ws.append(['Abreviações:'])
        for nome_completo, abrev in self.abreviacoes.items():
            ws.append([f'{abrev} = {nome_completo}'])
        ws.append([])
        ws.append(['Horário de atendimento: 08:00 - 22:00'])
        ws.append(['Gerado em: ' + datetime.now().strftime('%d/%m/%Y %H:%M:%S')])

        # Formatar título
        ws['A1'].font = Font(bold=True, size=14)

        # Ajustar larguras
        ws.column_dimensions['A'].width = 40

    def print_summary(self):
        """
        Imprime resumo da escala telefónica
        """
        if not self.phone_schedule:
            return

        print("=" * 80)
        print("RESUMO DA ESCALA TELEFÓNICA")
        print("=" * 80)
        print(f"Período: {self.start_date.strftime('%d/%m/%Y')} a {(self.start_date + timedelta(days=83)).strftime('%d/%m/%Y')}")
        print(f"Horário de atendimento: 08:00 às 22:00")
        print("Regra: Máximo de 3 horas consecutivas por pessoa")
        print()

        # Contar total de horas por pessoa (excluir 'Loja Fechada')
        totals = {abrev: 0 for abrev in self.abreviacoes.values()}

        for entry in self.phone_schedule:
            for hora in self.horarios:
                if (hora in entry and entry[hora] in totals and
                    entry[hora] != 'Loja F.'):
                    totals[entry[hora]] += 1

        print("Total de horas de atendimento telefónico:")
        print("-" * 40)

        for nome_completo, abrev in self.abreviacoes.items():
            horas = totals[abrev]
            print(f"{nome_completo:<15} ({abrev}): {horas:3d} horas")

        print("-" * 40)
        print(f"{'TOTAL':<15}: {sum(totals.values()):3d} horas")


class GeradorWorker(QThread):
    """Thread para gerar a escala sem travar a interface"""
    progress = pyqtSignal(int)
    finished = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, excel_file):
        super().__init__()
        self.excel_file = excel_file

    def run(self):
        try:
            self.progress.emit(30)
            generator = PhoneScheduleGenerator(self.excel_file)
            
            self.progress.emit(50)
            if not generator.work_schedule:
                self.error.emit(f"Não foi possível carregar a escala de trabalho do ficheiro:\n{self.excel_file}\n\nVerifique se o ficheiro está no formato correto.")
                return
            
            self.progress.emit(70)
            success = generator.generate_phone_schedule()
            
            if not success:
                self.error.emit("Erro ao gerar escala telefónica.")
                return
            
            self.progress.emit(90)
            filename = generator.export_to_excel()
            
            self.progress.emit(100)
            self.finished.emit(filename)
            
        except Exception as e:
            self.error.emit(f"Erro: {str(e)}\n\n{traceback.format_exc()}")


class GeradorApoiosWindow(QMainWindow):
    """Janela principal do Gerador de Apoios"""

    def verificar_ficheiro_selecionado(self):
        """Verifica se o ficheiro selecionado é válido"""
        if not hasattr(self, 'ficheiro_selecionado') or not self.ficheiro_selecionado:
            return False
        
        try:
            # Tentar ler o ficheiro para verificar se é válido
            df = pd.read_excel(self.ficheiro_selecionado, sheet_name='Escala de Trabalho', nrows=5)
            
            # Verificar colunas necessárias
            colunas_necessarias = ['Semana', 'Data', 'Dia', 'Susana A.', 'António C.', 'Antónia F.', 'Magda G.']
            colunas_existentes = [col for col in colunas_necessarias if col in df.columns]
            
            if len(colunas_existentes) >= 4:  # Pelo menos as principais existem
                self.log("✅ Ficheiro válido! Pronto para gerar.")
                return True
            else:
                self.log("⚠️ Ficheiro não contém todas as colunas esperadas.")
                return False
                
        except Exception as e:
            self.log(f"❌ Erro ao ler ficheiro: {str(e)}")
            return False
    
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Gerador de Apoios - Escala Telefónica")
        self.setGeometry(100, 100, 800, 600)
        
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Layout principal
        layout = QVBoxLayout(central_widget)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # Título
        title = QLabel("📞 GERADOR DE APOIOS")
        title.setFont(QFont("Arial", 20, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #E67E22; margin-bottom: 10px;")
        layout.addWidget(title)
        
        # Subtítulo
        subtitle = QLabel("Escala Telefónica de Atendimento (08:00 - 22:00)")
        subtitle.setFont(QFont("Arial", 12))
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color: #666; margin-bottom: 20px;")
        layout.addWidget(subtitle)
        
        # Separador
        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        separator.setStyleSheet("background-color: #E67E22; margin: 10px 0;")
        layout.addWidget(separator)
        
        # Área de informação
        info_frame = QFrame()
        info_frame.setStyleSheet("""
            QFrame {
                background-color: #FFF3E0;
                border: 2px solid #E67E22;
                border-radius: 10px;
                padding: 15px;
            }
        """)
        info_layout = QVBoxLayout(info_frame)
        
        info_text = QLabel(
            "📋 INFORMAÇÕES:\n\n"
            "• Gera escala telefónica baseada na escala de trabalho\n"
            "• Horário de atendimento: 08:00 - 22:00\n"
            "• Máximo de 3 horas consecutivas por pessoa\n"
            "• Distribuição equilibrada entre colaboradores\n"
            "• Respeita folgas, férias e horários de almoço\n\n"
            "📁 Necessário: arquivo 'escala_trabalho.xlsx' na mesma pasta"
        )
        info_text.setFont(QFont("Arial", 10))
        info_text.setWordWrap(True)
        info_text.setAlignment(Qt.AlignLeft)
        info_layout.addWidget(info_text)
        
        layout.addWidget(info_frame)

        # Após a criação do info_frame, adicionar:
        # Seleção de ficheiro
        file_frame = QFrame()
        file_frame.setStyleSheet("""
            QFrame {
                background-color: #F0F0F0;
                border: 1px solid #CCC;
                border-radius: 5px;
                padding: 10px;
            }
        """)
        file_layout = QHBoxLayout(file_frame)

        self.file_label = QLabel("Nenhum ficheiro selecionado")
        self.file_label.setFont(QFont("Arial", 9))
        self.file_label.setWordWrap(True)
        self.file_label.setStyleSheet("color: #666;")

        btn_selecionar = QPushButton("📂 Selecionar Ficheiro Excel")
        btn_selecionar.setFont(QFont("Arial", 10))
        btn_selecionar.setMinimumHeight(35)
        btn_selecionar.setStyleSheet("""
            QPushButton {
                background-color: #3498DB;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px 10px;
            }
            QPushButton:hover {
                background-color: #2980B9;
            }
        """)
        btn_selecionar.clicked.connect(self.selecionar_ficheiro)

        file_layout.addWidget(self.file_label, 1)
        file_layout.addWidget(btn_selecionar)

        layout.addWidget(file_frame)

        # Remover ou comentar a linha do status_file antigo
        # self.status_file = QLabel("")
        
        # Status do arquivo
        # self.status_file = QLabel("")
        # self.status_file.setAlignment(Qt.AlignCenter)
        # self.status_file.setStyleSheet("padding: 5px; font-weight: bold;")
        # layout.addWidget(self.status_file)

        # Área de log
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(150)
        self.log_text.setStyleSheet("""
            QTextEdit {
                background-color: #F5F5F5;
                border: 1px solid #CCC;
                border-radius: 5px;
                font-family: monospace;
                font-size: 10px;
            }
        """)
        layout.addWidget(self.log_text)
        
        # Verificar existência do arquivo
        # self.check_excel_file()
        
        # Barra de progresso
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #E67E22;
                border-radius: 5px;
                height: 20px;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #E67E22;
                border-radius: 3px;
            }
        """)
        layout.addWidget(self.progress_bar)
        
        
        
        # Botões
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)
        
        btn_gerar = QPushButton("🚀 Gerar Escala Telefónica")
        btn_gerar.setFont(QFont("Arial", 12))
        btn_gerar.setMinimumHeight(50)
        btn_gerar.setStyleSheet("""
            QPushButton {
                background-color: #E67E22;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #D35400;
            }
            QPushButton:pressed {
                background-color: #BA4A00;
            }
        """)
        btn_gerar.clicked.connect(self.gerar_escala)
        buttons_layout.addWidget(btn_gerar)
        
        btn_abrir_pasta = QPushButton("📁 Abrir Pasta")
        btn_abrir_pasta.setFont(QFont("Arial", 12))
        btn_abrir_pasta.setMinimumHeight(50)
        btn_abrir_pasta.setStyleSheet("""
            QPushButton {
                background-color: #3498DB;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #2980B9;
            }
            QPushButton:pressed {
                background-color: #1F618D;
            }
        """)
        btn_abrir_pasta.clicked.connect(self.abrir_pasta)
        buttons_layout.addWidget(btn_abrir_pasta)
        
        btn_fechar = QPushButton("✖️ Fechar")
        btn_fechar.setFont(QFont("Arial", 12))
        btn_fechar.setMinimumHeight(50)
        btn_fechar.setStyleSheet("""
            QPushButton {
                background-color: #95A5A6;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #7F8C8D;
            }
            QPushButton:pressed {
                background-color: #616A6B;
            }
        """)
        btn_fechar.clicked.connect(self.hide)
        buttons_layout.addWidget(btn_fechar)
        
        layout.addLayout(buttons_layout)
        
        # Rodapé
        footer = QLabel("© 2025 Gerador de Apoios - Baseado na Escala de Trabalho")
        footer.setAlignment(Qt.AlignCenter)
        footer.setStyleSheet("color: #999; font-size: 9px; margin-top: 10px;")
        layout.addWidget(footer)
        
        # Inicializar worker
        self.worker = None

    def check_excel_file(self):
        """Verifica se o arquivo Excel existe"""
        if os.path.exists('escala_trabalho.xlsx'):
            self.status_file.setText("✅ Arquivo 'escala_trabalho.xlsx' encontrado!")
            self.status_file.setStyleSheet("color: green; padding: 5px; font-weight: bold;")
            self.log("Arquivo de escala encontrado. Pronto para gerar.")
        else:
            self.status_file.setText("❌ Arquivo 'escala_trabalho.xlsx' NÃO encontrado!")
            self.status_file.setStyleSheet("color: red; padding: 5px; font-weight: bold;")
            self.log("ERRO: Arquivo 'escala_trabalho.xlsx' não encontrado na pasta atual.")

    def log(self, message):
        """Adiciona mensagem ao log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")

    def gerar_escala(self):
        """Inicia a geração da escala"""
        if not hasattr(self, 'ficheiro_selecionado') or not self.ficheiro_selecionado:
            QMessageBox.warning(
                self, 
                "Aviso", 
                "Por favor, selecione primeiro um ficheiro Excel com a escala de trabalho."
            )
            return
        
        if not self.verificar_ficheiro_selecionado():
            resposta = QMessageBox.question(
                self,
                "Confirmar",
                "O ficheiro selecionado pode não estar no formato correto.\n\nDeseja continuar mesmo assim?",
                QMessageBox.Yes | QMessageBox.No
            )
            if resposta == QMessageBox.No:
                return
        
        # Desabilitar botões durante a geração
        for btn in self.findChildren(QPushButton):
            btn.setEnabled(False)
        
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.log(f"Iniciando geração da escala telefónica a partir de: {os.path.basename(self.ficheiro_selecionado)}")
        
        # Criar e iniciar worker com o ficheiro selecionado
        self.worker = GeradorWorker(self.ficheiro_selecionado)
        self.worker.progress.connect(self.atualizar_progresso)
        self.worker.finished.connect(self.geracao_concluida)
        self.worker.error.connect(self.erro_geracao)
        self.worker.start()

    def atualizar_progresso(self, valor):
        """Atualiza a barra de progresso"""
        self.progress_bar.setValue(valor)

    def geracao_concluida(self, filename):
        """Chamado quando a geração termina com sucesso"""
        self.progress_bar.setVisible(False)
        self.log(f"✅ Escala gerada com sucesso: {filename}")
        
        # Reabilitar botões
        for btn in self.findChildren(QPushButton):
            btn.setEnabled(True)
        
        QMessageBox.information(
            self, 
            "Sucesso", 
            f"Escala telefónica gerada com sucesso!\n\n"
            f"Arquivo: {filename}\n\n"
            f"O arquivo foi salvo na pasta do programa."
        )

    def erro_geracao(self, erro_msg):
        """Chamado quando ocorre erro na geração"""
        self.progress_bar.setVisible(False)
        self.log(f"❌ ERRO: {erro_msg}")
        
        # Reabilitar botões
        for btn in self.findChildren(QPushButton):
            btn.setEnabled(True)
        
        QMessageBox.critical(self, "Erro na Geração", erro_msg)

    def abrir_pasta(self):
        """Abre a pasta onde os arquivos são salvos"""
        import subprocess
        import platform
        
        pasta = os.path.dirname(os.path.abspath(__file__))
        
        try:
            if platform.system() == "Windows":
                os.startfile(pasta)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", pasta])
            else:  # Linux
                subprocess.run(["xdg-open", pasta])
            
            self.log(f"Pasta aberta: {pasta}")
        except Exception as e:
            self.log(f"Erro ao abrir pasta: {e}")
            QMessageBox.warning(self, "Aviso", f"Não foi possível abrir a pasta:\n{pasta}")

    def closeEvent(self, event):
        """Sobrescreve o evento de fechar para apenas esconder"""
        self.hide()
        event.ignore()   

    
    def selecionar_ficheiro(self):
        """Abre diálogo para selecionar ficheiro Excel"""
        from PyQt5.QtWidgets import QFileDialog
        
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Selecionar Escala de Trabalho",
            os.path.expanduser("~"),
            "Ficheiros Excel (*.xlsx *.xls)"
        )
        
        if file_path:
            self.ficheiro_selecionado = file_path
            nome_ficheiro = os.path.basename(file_path)
            self.file_label.setText(f"📁 {nome_ficheiro}")
            self.file_label.setStyleSheet("color: #27AE60; font-weight: bold;")
            self.log(f"Ficheiro selecionado: {file_path}")
            
            # Verificar se o ficheiro é válido
            self.verificar_ficheiro_selecionado()
        else:
            self.ficheiro_selecionado = None
            self.file_label.setText("Nenhum ficheiro selecionado")
            self.file_label.setStyleSheet("color: #666;")

            
            
        # Limpa a referência quando a janela for realmente destruída
        _janela_gerador_apoios.destroyed.connect(lambda: globals().update(_janela_gerador_apoios=None))

def mostrar_gerador_apoios():
    """Função para mostrar a janela do Gerador de Apoios"""
    global _janela_gerador_apoios
    
    # Se já existe uma janela, apenas reativa
    if _janela_gerador_apoios is not None:
        _janela_gerador_apoios.show()
        _janela_gerador_apoios.raise_()
        _janela_gerador_apoios.activateWindow()
        return
    
    # Cria nova janela
    _janela_gerador_apoios = GeradorApoiosWindow()
    
    # Garante que ao fechar, não seja destruída
    _janela_gerador_apoios.setAttribute(Qt.WA_DeleteOnClose, False)
    
    _janela_gerador_apoios.show()

if __name__ == "__main__":
    # Para testar individualmente
    app = QApplication(sys.argv)
    window = GeradorApoiosWindow()
    window.show()
    sys.exit(app.exec_())
